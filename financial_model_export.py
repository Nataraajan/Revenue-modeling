"""
Generates a traditional financial-model-format Excel workbook covering
MULTIPLE pods (segments) at once — a real company-wide model, not a
single team's numbers — since a workbook shared with bankers/investors/
corp dev is expected to represent the whole business.

Structure:
  - Assumptions: one column per pod, shared row labels (easy side-by-side
    comparison of segment economics).
  - Per pod: "{Pod} Capacity" and "{Pod} Revenue" sheets — same proven
    cohort-waterfall logic as the single-pod version, just namespaced.
  - Summary: one Year1/Year2/... column PER POD, plus a "Total Company"
    column per year that SUMS across pods — the standard way a real
    multi-segment financial model rolls up (segment columns + total column).

Same known limitations as before: no PS fees, no seasonality, no Existing
Book overlay, full years only in the Summary.
"""

import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

BLUE = Font(color="0000FF")
BLACK = Font(color="000000")
GREEN = Font(color="008000")
BOLD = Font(bold=True)
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CURRENCY_FMT = '[$$-409]#,##0;([$$-409]#,##0);"-"'
PCT_FMT = '0.0%'
NUM_FMT = '#,##0'


def _phased_hire_months(count: int, cadence: int) -> list:
    return [i * cadence for i in range(count)]


def _safe_sheet_name(pod_name: str, suffix: str, used_names: set) -> str:
    base = pod_name[:31 - len(suffix) - 1]
    name = f"{base} {suffix}"
    i = 2
    while name in used_names:
        base2 = pod_name[:31 - len(suffix) - 3]
        name = f"{base2}{i} {suffix}"
        i += 1
    used_names.add(name)
    return name


def _build_pod_assumptions_column(ws, cfg: dict, col: int, num_months: int) -> dict:
    term = int(cfg["contract_term"])
    lag = int(cfg["implementation_lag"])
    col_letter = get_column_letter(col)

    values = [
        num_months, term, lag, None,
        cfg["ae_base"], cfg["ae_variable"], cfg["ae_quota"],
        cfg["bdr_base"], cfg["bdr_variable"], cfg["bdr_monthly_sql_quota"], None,
        cfg["marketing_sqls"], cfg["ae_self_sourced"], None,
        cfg["avg_deal_size"], cfg["win_marketing"], cfg["win_bdr"], cfg["win_self"],
        cfg["execution_efficiency"], None,
        cfg["churn_rate"], cfg["expansion_rate"], cfg["contraction_rate"],
    ]
    r = 4
    ref_rows = {}
    labels_in_order = ["horizon", "term", "lag", None, "ae_base", "ae_variable", "ae_quota",
                        "bdr_base", "bdr_variable", "bdr_sql_quota", None,
                        "marketing_sqls", "self_sourced", None,
                        "deal_size", "win_mkt", "win_bdr", "win_self", "exec_eff", None,
                        "churn", "expansion", "contraction"]
    for label, value in zip(labels_in_order, values):
        if label:
            c = ws.cell(row=r, column=col, value=value)
            c.font = BLUE
            _PERCENT_FIELDS = {"win_mkt", "win_bdr", "win_self", "churn", "expansion", "contraction", "exec_eff"}
            if label in _PERCENT_FIELDS:
                c.number_format = PCT_FMT
            elif label in ("ae_base", "ae_variable", "ae_quota", "bdr_base", "bdr_variable", "deal_size"):
                c.number_format = CURRENCY_FMT
            else:
                c.number_format = NUM_FMT
            ref_rows[label] = f"Assumptions!{col_letter}${r}"
        r += 1
    return ref_rows


def generate_multi_pod_workbook_bytes(cfg_list: list, num_months: int) -> bytes:
    wb = openpyxl.Workbook()
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a["A1"] = "Revenue Architecture Model — Multi-Segment"
    ws_a["A1"].font = Font(bold=True, size=14)
    ws_a["A2"] = "Blue = input. Black = formula. Green = link to another sheet. One column per segment."
    ws_a["A2"].font = Font(italic=True, size=9)

    row_labels = [
        "Horizon (months)", "Contract term (months)", "Implementation lag (months)", "",
        "AE — annual base ($)", "AE — annual variable @ 100% ($)", "AE — annual quota ($)",
        "BDR — annual base ($)", "BDR — annual variable @ 100% ($)", "BDR — monthly SQL quota", "",
        "Marketing SQLs/month", "AE self-sourced SQLs/month", "",
        "Avg deal size — TCV ($)", "Win rate — marketing-sourced", "Win rate — BDR-sourced",
        "Win rate — AE self-sourced", "Execution efficiency", "",
        "Renewal — annual gross churn rate", "Renewal — annual expansion rate", "Renewal — annual contraction rate",
    ]
    for i, label in enumerate(row_labels):
        ws_a.cell(row=4 + i, column=1, value=label)
    ws_a.column_dimensions["A"].width = 34

    used_sheet_names = {"Assumptions"}
    pod_data = []

    for pi, cfg in enumerate(cfg_list):
        col = 2 + pi
        ws_a.cell(row=3, column=col, value=cfg["pod_name"]).font = Font(bold=True)
        ws_a.column_dimensions[get_column_letter(col)].width = 14
        ref = _build_pod_assumptions_column(ws_a, cfg, col, num_months)

        cap_sheet_name = _safe_sheet_name(cfg["pod_name"], "Capacity", used_sheet_names)
        rr_sheet_name = _safe_sheet_name(cfg["pod_name"], "Revenue", used_sheet_names)

        info = _build_capacity_sheet(wb, cfg, ref, num_months, cap_sheet_name)
        info2 = _build_revenue_sheet(wb, cfg, ref, num_months, rr_sheet_name, cap_sheet_name, info["bookings_row"])
        pod_data.append({"pod_name": cfg["pod_name"], "rr_sheet": rr_sheet_name, **info2})

    _build_company_summary(wb, pod_data, num_months)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_capacity_sheet(wb, cfg, ref, num_months, sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"{cfg['pod_name']} — Capacity & Bookings"
    ws["A1"].font = Font(bold=True, size=14)
    CAP_FIRST_COL = 5

    ws.cell(row=3, column=4, value="Month →").font = BOLD
    for m in range(1, num_months + 1):
        c = ws.cell(row=3, column=CAP_FIRST_COL + m - 1, value=m)
        c.font = BOLD
        c.fill = HEADER_FILL

    ae_hires_0idx = cfg.get("ae_hire_months") or _phased_hire_months(int(cfg["num_aes"]), int(cfg["hiring_cadence"]))
    bdr_hires_0idx = cfg.get("bdr_hire_months") or _phased_hire_months(int(cfg["num_bdrs"]), int(cfg["hiring_cadence"]))
    ae_hires = [h + 1 for h in ae_hires_0idx]
    bdr_hires = [h + 1 for h in bdr_hires_0idx]
    ae_roster = [(f"AE-existing-{i+1}", 1, "Instant") for i in range(int(cfg["num_existing_aes"]))]
    ae_roster += [(f"AE-new-{i+1}", ae_hires[i], "Standard") for i in range(int(cfg["num_aes"]))]
    bdr_roster = [(f"BDR-existing-{i+1}", 1, "Instant") for i in range(int(cfg["num_existing_bdrs"]))]
    bdr_roster += [(f"BDR-new-{i+1}", bdr_hires[i], "Standard") for i in range(int(cfg["num_bdrs"]))]

    def _write_ramp_row(row):
        for m in range(1, num_months + 1):
            col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
            month_cell = f"{col_letter}$3"
            formula = (
                f'=IF({month_cell}<$B{row},0,'
                f'IF($C{row}="Instant",1,'
                f'IF({month_cell}-$B{row}>=2,1,'
                f'IF({month_cell}-$B{row}=1,0.66,0.33))))'
            )
            c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1, value=formula)
            c.font = BLACK
            c.number_format = PCT_FMT

    row = 5
    ws.cell(row=row, column=1, value="AE Roster").font = BOLD
    row += 1
    ae_ramp_rows = []
    for name, hire_month, ramp_type in ae_roster:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=hire_month).font = BLUE
        ws.cell(row=row, column=3, value=ramp_type).font = BLUE
        _write_ramp_row(row)
        ae_ramp_rows.append(row)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="AE $ Capacity by Rep").font = BOLD
    row += 1
    ae_capacity_rows = []
    for i, _ in enumerate(ae_roster):
        ramp_row = ae_ramp_rows[i]
        ws.cell(row=row, column=1, value=ae_roster[i][0])
        for m in range(1, num_months + 1):
            col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
            c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1, value=f'={col_letter}{ramp_row}*({ref["ae_quota"]}/12)')
            c.font = BLACK
            c.number_format = CURRENCY_FMT
        ae_capacity_rows.append(row)
        row += 1

    total_ae_capacity_row = row
    ws.cell(row=row, column=1, value="Total AE $ Capacity").font = BOLD
    for m in range(1, num_months + 1):
        col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
        refs = "+".join(f"{col_letter}{rr}" for rr in ae_capacity_rows) if ae_capacity_rows else "0"
        c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1, value=f"={refs}")
        c.font = BOLD
        c.number_format = CURRENCY_FMT
    row += 2

    ws.cell(row=row, column=1, value="AE Self-Sourced SQLs by Rep").font = BOLD
    row += 1
    self_sourced_rows = []
    for i, _ in enumerate(ae_roster):
        ramp_row = ae_ramp_rows[i]
        ws.cell(row=row, column=1, value=ae_roster[i][0])
        for m in range(1, num_months + 1):
            col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
            c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1, value=f'={col_letter}{ramp_row}*{ref["self_sourced"]}')
            c.font = BLACK
            c.number_format = NUM_FMT
        self_sourced_rows.append(row)
        row += 1

    total_self_sourced_row = row
    ws.cell(row=row, column=1, value="Total AE Self-Sourced SQLs").font = BOLD
    for m in range(1, num_months + 1):
        col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
        refs = "+".join(f"{col_letter}{rr}" for rr in self_sourced_rows) if self_sourced_rows else "0"
        c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1, value=f"={refs}")
        c.font = BOLD
        c.number_format = NUM_FMT
    row += 2

    ws.cell(row=row, column=1, value="BDR Roster").font = BOLD
    row += 1
    bdr_ramp_rows = []
    for name, hire_month, ramp_type in bdr_roster:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=hire_month).font = BLUE
        ws.cell(row=row, column=3, value=ramp_type).font = BLUE
        _write_ramp_row(row)
        bdr_ramp_rows.append(row)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="BDR SQLs by Rep").font = BOLD
    row += 1
    bdr_sql_rows = []
    for i, _ in enumerate(bdr_roster):
        ramp_row = bdr_ramp_rows[i]
        ws.cell(row=row, column=1, value=bdr_roster[i][0])
        for m in range(1, num_months + 1):
            col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
            c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1, value=f'={col_letter}{ramp_row}*{ref["bdr_sql_quota"]}')
            c.font = BLACK
            c.number_format = NUM_FMT
        bdr_sql_rows.append(row)
        row += 1

    total_bdr_sql_row = row
    ws.cell(row=row, column=1, value="Total BDR SQLs").font = BOLD
    for m in range(1, num_months + 1):
        col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
        refs = "+".join(f"{col_letter}{rr}" for rr in bdr_sql_rows) if bdr_sql_rows else "0"
        c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1, value=f"={refs}")
        c.font = BOLD
        c.number_format = NUM_FMT
    row += 2

    marketing_sql_row = row
    ws.cell(row=row, column=1, value="Marketing SQLs").font = BOLD
    for m in range(1, num_months + 1):
        col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
        c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1, value=f'={ref["marketing_sqls"]}')
        c.font = BLACK
        c.number_format = NUM_FMT
    row += 2

    demand_row = row
    ws.cell(row=row, column=1, value="Demand-Constrained Bookings ($)").font = BOLD
    for m in range(1, num_months + 1):
        col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
        mkt = f"{col_letter}{marketing_sql_row}*{ref['win_mkt']}*{ref['deal_size']}"
        bdr = f"{col_letter}{total_bdr_sql_row}*{ref['win_bdr']}*{ref['deal_size']}"
        slf = f"{col_letter}{total_self_sourced_row}*{ref['win_self']}*{ref['deal_size']}"
        c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1, value=f"={mkt}+{bdr}+{slf}")
        c.font = BLACK
        c.number_format = CURRENCY_FMT
    row += 1

    bookings_row = row
    ws.cell(row=row, column=1, value="Actual Bookings ($ TCV)").font = Font(bold=True, size=11)
    for m in range(1, num_months + 1):
        col_letter = get_column_letter(CAP_FIRST_COL + m - 1)
        c = ws.cell(row=row, column=CAP_FIRST_COL + m - 1,
                     value=f'=MIN({col_letter}{total_ae_capacity_row},{col_letter}{demand_row})*{ref["exec_eff"]}')
        c.font = Font(bold=True)
        c.number_format = CURRENCY_FMT

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    return {"bookings_row": bookings_row, "cap_first_col": CAP_FIRST_COL}


def _build_revenue_sheet(wb, cfg, ref, num_months, sheet_name, cap_sheet_name, bookings_row):
    term = int(cfg["contract_term"])
    lag = int(cfg["implementation_lag"])
    num_generations = 1
    g = 1
    while (1 + lag + g * term) <= num_months:
        num_generations += 1
        g += 1

    ws = wb.create_sheet(sheet_name)
    ws["A1"] = f"{cfg['pod_name']} — Revenue Recognition"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"{num_generations} generation(s) fit in this {num_months}-month horizon at a {term}-month term."
    ws["A2"].font = Font(italic=True, size=9)

    RR_FIRST_COL = 11
    ws.cell(row=4, column=4, value="Month →").font = BOLD
    for m in range(1, num_months + 1):
        c = ws.cell(row=4, column=RR_FIRST_COL + m - 1, value=m)
        c.font = BOLD
        c.fill = HEADER_FILL

    headers = ["Cohort", "Signed Mo.", "Go-Live Mo.", "Rec. End Mo.", "Cohort TCV", "Cohort ARR",
               "Churned $", "Retained $", "Expansion $", "Contraction $"]
    for i, h in enumerate(headers):
        ws.cell(row=5, column=1 + i, value=h).font = BOLD

    row = 6
    all_gen_rows = []
    prev_gen_row_lookup = {}

    for gen in range(num_generations):
        if gen > 0:
            row += 1
            ws.cell(row=row, column=1, value=f"Renewal Generation {gen}").font = BOLD
            row += 1
        gen_row_lookup = {}
        for m in range(1, num_months + 1):
            label = f"Gen {gen} — Month {m}" if gen == 0 else f"Gen {gen} — renews Month {m}"
            ws.cell(row=row, column=1, value=label)

            if gen == 0:
                ws.cell(row=row, column=2, value=m).font = BLACK
                ws.cell(row=row, column=3, value=f"=B{row}+{ref['lag']}").font = BLACK
                tcv_cell = ws.cell(row=row, column=5,
                                    value=f"='{cap_sheet_name}'!{get_column_letter(5 + m - 1)}{bookings_row}")
                tcv_cell.font = GREEN
                tcv_cell.number_format = CURRENCY_FMT
            else:
                prev_row = prev_gen_row_lookup.get((gen - 1, m))
                if prev_row is None:
                    row += 1
                    continue
                ws.cell(row=row, column=2, value=f"=D{prev_row}+1").font = BLACK
                ws.cell(row=row, column=3, value=f"=B{row}").font = BLACK
                tcv_cell = ws.cell(row=row, column=5, value=f"=(H{prev_row}+I{prev_row}-J{prev_row})*({ref['term']}/12)")
                tcv_cell.font = BLACK
                tcv_cell.number_format = CURRENCY_FMT

            ws.cell(row=row, column=4, value=f"=C{row}+{ref['term']}-1").font = BLACK
            arr_cell = ws.cell(row=row, column=6, value=f"=E{row}*(12/{ref['term']})")
            arr_cell.font = BLACK
            arr_cell.number_format = CURRENCY_FMT

            if gen >= 1 or gen < num_generations - 1:
                ws.cell(row=row, column=7, value=f"=F{row}*{ref['churn']}").font = BLACK
                ws.cell(row=row, column=7).number_format = CURRENCY_FMT
                ws.cell(row=row, column=8, value=f"=F{row}-G{row}").font = BLACK
                ws.cell(row=row, column=8).number_format = CURRENCY_FMT
                ws.cell(row=row, column=9, value=f"=H{row}*{ref['expansion']}").font = BLACK
                ws.cell(row=row, column=9).number_format = CURRENCY_FMT
                ws.cell(row=row, column=10, value=f"=H{row}*{ref['contraction']}").font = BLACK
                ws.cell(row=row, column=10).number_format = CURRENCY_FMT

            for mm in range(1, num_months + 1):
                col_letter = get_column_letter(RR_FIRST_COL + mm - 1)
                month_header_cell = f"{col_letter}$4"
                formula = f'=IF(AND({month_header_cell}>=$C{row},{month_header_cell}<=$D{row}),$E{row}/{ref["term"]},0)'
                c = ws.cell(row=row, column=RR_FIRST_COL + mm - 1, value=formula)
                c.font = BLACK
                c.number_format = CURRENCY_FMT

            gen_row_lookup[m] = row
            all_gen_rows.append({"gen": gen, "month": m, "row": row})
            row += 1
        prev_gen_row_lookup = {(gen, m): r for m, r in gen_row_lookup.items()}

    row += 1
    total_rev_row = row
    ws.cell(row=row, column=1, value="Total Recognized Revenue").font = Font(bold=True, size=11)
    all_data_rows = [d["row"] for d in all_gen_rows]
    for mm in range(1, num_months + 1):
        col_letter = get_column_letter(RR_FIRST_COL + mm - 1)
        refs = "+".join(f"{col_letter}{r}" for r in all_data_rows)
        c = ws.cell(row=row, column=RR_FIRST_COL + mm - 1, value=f"={refs}")
        c.font = Font(bold=True)
        c.number_format = CURRENCY_FMT
    row += 1

    live_arr_row = row
    ws.cell(row=row, column=1, value="Live ARR").font = Font(bold=True, size=11)
    for mm in range(1, num_months + 1):
        col_letter = get_column_letter(RR_FIRST_COL + mm - 1)
        month_header_cell = f"{col_letter}$4"
        terms = [f'IF(AND({month_header_cell}>=$C{r},{month_header_cell}<=$D{r}),$F{r},0)' for r in all_data_rows]
        formula = "=" + "+".join(terms) if terms else "=0"
        c = ws.cell(row=row, column=RR_FIRST_COL + mm - 1, value=formula)
        c.font = Font(bold=True)
        c.number_format = CURRENCY_FMT

    ws.column_dimensions["A"].width = 26
    for cl in ["E", "F", "G", "H", "I", "J"]:
        ws.column_dimensions[cl].width = 13

    gen0_rows = sorted([d["row"] for d in all_gen_rows if d["gen"] == 0])
    gen1_rows = sorted([d["row"] for d in all_gen_rows if d["gen"] == 1]) if num_generations > 1 else []

    return {
        "total_rev_row": total_rev_row, "live_arr_row": live_arr_row,
        "gen0_rows": gen0_rows, "gen1_rows": gen1_rows, "rr_first_col": RR_FIRST_COL,
    }


def _build_company_summary(wb, pod_data, num_months):
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Company Summary — Segment + Total"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "'New ARR' sums Gen0 cohorts by go-live month. Total Company columns sum across all segments."
    ws["A2"].font = Font(italic=True, size=9)

    num_full_years = num_months // 12
    years = [(y + 1, y * 12 + 1, y * 12 + 12) for y in range(num_full_years)]

    labels = ["Beginning ARR", "New ARR (went live)", "Expansion ARR", "Contraction ARR", "Churned ARR",
              "Other / Boundary Effect", "Ending ARR", "", "Net Revenue Retention (NRR)", "Gross $ Churn Rate",
              "", "Total Revenue Recognized"]
    metric_row = {}
    r = 5
    for label in labels:
        ws.cell(row=r, column=1, value=label).font = BOLD if label and "Other" not in label else BLACK
        if label:
            metric_row[label] = r
        r += 1

    col = 2
    pod_col_blocks = {}
    for pod in pod_data:
        cols_this_pod = []
        for yr, s, e in years:
            ws.cell(row=3, column=col, value=pod["pod_name"]).font = Font(bold=True)
            c = ws.cell(row=4, column=col, value=f"Year {yr}")
            c.font = BOLD
            c.fill = HEADER_FILL
            cols_this_pod.append((yr, col))
            col += 1
        pod_col_blocks[pod["pod_name"]] = cols_this_pod

    total_cols = []
    for yr, s, e in years:
        ws.cell(row=3, column=col, value="TOTAL COMPANY").font = Font(bold=True)
        c = ws.cell(row=4, column=col, value=f"Year {yr}")
        c.font = BOLD
        c.fill = TOTAL_FILL
        total_cols.append((yr, col))
        col += 1

    def rr_col(first_col, month):
        return get_column_letter(first_col + month - 1)

    for pod in pod_data:
        rr = f"'{pod['rr_sheet']}'!"
        gen0_rows, gen1_rows = pod["gen0_rows"], pod["gen1_rows"]
        rr_first_col = pod["rr_first_col"]

        for (yr, col) in pod_col_blocks[pod["pod_name"]]:
            _, s, e = [y for y in years if y[0] == yr][0]
            col_letter = get_column_letter(col)

            beg_cell = ws.cell(row=metric_row["Beginning ARR"], column=col)
            if yr == 1:
                beg_cell.value = 0
                beg_cell.font = BLUE
            else:
                prev_col_letter = get_column_letter(col - 1)
                beg_cell.value = f"={prev_col_letter}{metric_row['Ending ARR']}"
                beg_cell.font = BLACK
            beg_cell.number_format = CURRENCY_FMT

            end_cell = ws.cell(row=metric_row["Ending ARR"], column=col,
                                value=f"={rr}{rr_col(rr_first_col, e)}{pod['live_arr_row']}")
            end_cell.font = GREEN
            end_cell.number_format = CURRENCY_FMT

            if gen0_rows:
                new_arr_cell = ws.cell(
                    row=metric_row["New ARR (went live)"], column=col,
                    value=(f"=SUMIFS({rr}F{gen0_rows[0]}:F{gen0_rows[-1]},"
                           f"{rr}C{gen0_rows[0]}:C{gen0_rows[-1]},\">=\"&{s},"
                           f"{rr}C{gen0_rows[0]}:C{gen0_rows[-1]},\"<=\"&{e})")
                )
            else:
                new_arr_cell = ws.cell(row=metric_row["New ARR (went live)"], column=col, value=0)
            new_arr_cell.font = GREEN
            new_arr_cell.number_format = CURRENCY_FMT

            if gen1_rows:
                for label, gcol in [("Expansion ARR", "I"), ("Contraction ARR", "J"), ("Churned ARR", "G")]:
                    c = ws.cell(
                        row=metric_row[label], column=col,
                        value=(f"=SUMIFS({rr}{gcol}{gen1_rows[0]}:{gcol}{gen1_rows[-1]},"
                               f"{rr}D{gen0_rows[0]}:D{gen0_rows[-1]},\">=\"&{s},"
                               f"{rr}D{gen0_rows[0]}:D{gen0_rows[-1]},\"<=\"&{e})")
                    )
                    c.font = GREEN
                    c.number_format = CURRENCY_FMT
            else:
                for label in ["Expansion ARR", "Contraction ARR", "Churned ARR"]:
                    c = ws.cell(row=metric_row[label], column=col, value=0)
                    c.font = BLACK
                    c.number_format = CURRENCY_FMT

            beg_ref = f"{col_letter}{metric_row['Beginning ARR']}"
            new_ref = f"{col_letter}{metric_row['New ARR (went live)']}"
            exp_ref = f"{col_letter}{metric_row['Expansion ARR']}"
            contr_ref = f"{col_letter}{metric_row['Contraction ARR']}"
            churn_ref = f"{col_letter}{metric_row['Churned ARR']}"
            end_ref = f"{col_letter}{metric_row['Ending ARR']}"
            other_cell = ws.cell(row=metric_row["Other / Boundary Effect"], column=col,
                                  value=f"={end_ref}-({beg_ref}+{new_ref}+{exp_ref}-{contr_ref}-{churn_ref})")
            other_cell.font = BLACK
            other_cell.number_format = CURRENCY_FMT

            nrr_cell = ws.cell(row=metric_row["Net Revenue Retention (NRR)"], column=col,
                                value=f'=IFERROR(({beg_ref}+{exp_ref}-{contr_ref}-{churn_ref})/{beg_ref},"N/A")')
            nrr_cell.font = BLACK
            nrr_cell.number_format = PCT_FMT

            churnrate_cell = ws.cell(row=metric_row["Gross $ Churn Rate"], column=col,
                                      value=f'=IFERROR({churn_ref}/{beg_ref},"N/A")')
            churnrate_cell.font = BLACK
            churnrate_cell.number_format = PCT_FMT

            rev_cell = ws.cell(row=metric_row["Total Revenue Recognized"], column=col,
                                value=f"=SUM({rr}{rr_col(rr_first_col, s)}{pod['total_rev_row']}:{rr_col(rr_first_col, e)}{pod['total_rev_row']})")
            rev_cell.font = GREEN
            rev_cell.number_format = CURRENCY_FMT

    for (yr, tcol) in total_cols:
        tcol_letter = get_column_letter(tcol)
        pod_cols_this_year = [dict(pod_col_blocks[p["pod_name"]])[yr] for p in pod_data]
        for label in ["Beginning ARR", "New ARR (went live)", "Expansion ARR", "Contraction ARR",
                      "Churned ARR", "Other / Boundary Effect", "Ending ARR", "Total Revenue Recognized"]:
            r = metric_row[label]
            refs = "+".join(f"{get_column_letter(c)}{r}" for c in pod_cols_this_year)
            cell = ws.cell(row=r, column=tcol, value=f"={refs}")
            cell.font = Font(bold=True)
            cell.number_format = CURRENCY_FMT
            cell.fill = TOTAL_FILL

        beg_ref = f"{tcol_letter}{metric_row['Beginning ARR']}"
        exp_ref = f"{tcol_letter}{metric_row['Expansion ARR']}"
        contr_ref = f"{tcol_letter}{metric_row['Contraction ARR']}"
        churn_ref = f"{tcol_letter}{metric_row['Churned ARR']}"
        nrr_cell = ws.cell(row=metric_row["Net Revenue Retention (NRR)"], column=tcol,
                            value=f'=IFERROR(({beg_ref}+{exp_ref}-{contr_ref}-{churn_ref})/{beg_ref},"N/A")')
        nrr_cell.font = Font(bold=True)
        nrr_cell.number_format = PCT_FMT
        nrr_cell.fill = TOTAL_FILL

        churnrate_cell = ws.cell(row=metric_row["Gross $ Churn Rate"], column=tcol,
                                  value=f'=IFERROR({churn_ref}/{beg_ref},"N/A")')
        churnrate_cell.font = Font(bold=True)
        churnrate_cell.number_format = PCT_FMT
        churnrate_cell.fill = TOTAL_FILL

    ws.column_dimensions["A"].width = 30
